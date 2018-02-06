#!/usr/bin/env python

'''
---AUTHOR---
Name: Matt Cross
Email: routeallthings@gmail.com

---PREREQ---
INSTALL netmiko (pip install netmiko)
INSTALL textfsm (pip install textfsm)
INSTALL openpyxl (pip install openpyxl)
INSTALL fileinput (pip install fileinput)
INSTALL xlhelper (python -m pip install git+git://github.com/routeallthings/xlhelper.git)
'''

#Module Imports (Native)
import re
import getpass
import os
import unicodedata
import csv
import threading
import time
import sys
from datetime import datetime

#Module Imports (Non-Native)
try:
	import netmiko
	from netmiko import ConnectHandler
except ImportError:
	netmikoinstallstatus = fullpath = raw_input ('Netmiko module is missing, would you like to automatically install? (Y/N): ')
	if "Y" in netmikoinstallstatus.upper() or "YES" in netmikoinstallstatus.upper():
		os.system('python -m pip install netmiko')
		import netmiko
		from netmiko import ConnectHandler
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of netmiko. Please install manually and retry"
		sys.exit()
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('openpyxl module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install openpyxl')
		from openpyxl import load_workbook
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of Pandas. Please install manually and retry'
		sys.exit()
#
try:
	import fileinput
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('FileInput module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install FileInput')
		import FileInput
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of FileInput. Please install manually and retry'
		sys.exit()
# Darth-Veitcher Module https://github.com/darth-veitcher/xlhelper		
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import OrderedDict
try:
	import xlhelper
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('xlhelper module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install git+git://github.com/routeallthings/xlhelper.git')
		import xlhelper
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of xlhelper. Please install manually and retry'
		sys.exit()
#######################################
#Functions
def GetVLANs(xlsxfile,vlanfolder):
	# Build initial global lists
	for devices in xlhelper.sheet_to_dict(xlsxfile,'Device IPs'):
		devicelist.append(devices)
	for vlans in xlhelper.sheet_to_dict(xlsxfile,'VLAN List'):
		vlanlist.append(vlans)
	# Parse through each individual list
	for device in devicelist:
		devicehostname = device.get('Hostname').encode('utf-8')
		devicepath = vlanfolder + '\\' + devicehostname + '-vlans.txt'
		deviceopenw = open(devicepath,'w+')
		nextline = '\n'
		for assignment in xlhelper.sheet_to_dict(xlsxfile,devicehostname):
			try:
				portname = assignment.get('Port').encode('utf-8')
			except:
				portname = assignment.get('Port')
			if portname == None:
				portname = ''
			try:
				portvlan = assignment.get('VLAN').encode('utf-8')
			except:
				portvlan = assignment.get('VLAN')
			if portvlan == None:
				portvlan = ''
			try:
				porttemplate = assignment.get('Template').encode('utf-8')
			except:
				porttemplate = assignment.get('Template')
			if porttemplate == None:
				porttemplate = ''
			try:
				portdesc = assignment.get('Description').encode('utf-8')
			except:
				portdesc = assignment.get('Description')
			if portdesc == None:
				portdesc = ''
			portall = portname + ',' + portvlan + ',' + porttemplate + ',' + portdesc
			deviceopenw.write(portall)
			deviceopenw.write(nextline)
		deviceopenw.close()

def UpdateVLANs(device,vlanfolder,vlanlist):
	devicehost = device.get('Hostname').encode('utf-8')
	deviceip = device.get('IP').encode('utf-8')
	devicevendor = device.get('Vendor').encode('utf-8')
	devicetype = device.get('Type').encode('utf-8')
	devicetype = devicevendor.lower() + "_" + devicetype.lower()
	# Build Assignment List
	assignmentfile = vlanfolder + '\\' + devicehost + '-vlans.txt'
	assignmentlist = []
	assignmentfileo = open(assignmentfile,'r')
	assignmentdata = assignmentfileo.readlines()
	assignmentfileo.close()
	for line in assignmentdata:
		linedata = line.lstrip()
		linedata = linedata.strip('\n')
		linedata = linedata.split(',')
		assignmentlist.append(linedata)
	# Build Configuration Set
	cmdlist = []
	for line in assignmentlist:
		try:
			lineport = line[0]
		except:
			lineport = ''
		try:
			linevlan = line[1]
			for vlanl in vlanlist:
				try:
					vlanname = str(vlanl.get('VLAN Name').encode('utf-8'))
				except:
					vlanname = str(vlanl.get('VLAN Name'))
				if linevlan == vlanname:
					linevlan = str(vlanl.get('VLAN #'))
		except:
			linevlan = ''
		try:
			linetemplate = line[2]
		except:
			linetemplate = ''
		try:
			linedesc = line[3]
		except:
			linedesc = ''
		if not lineport == '':
			intcmd = 'interface ' + lineport
			cmdlist.append(intcmd)
		if not linevlan == '':
			intvlan = 'switchport access vlan ' + linevlan
			cmdlist.append(intvlan)
		if not linetemplate == '':
			inttemplate = 'source template ' + linetemplate
			cmdlist.append(inttemplate)
		if not linedesc == '':
			intdesc = 'description ' + linedesc
			cmdlist.append(intdesc)
	# Start Connection
	try:
		sshnet_connect = ConnectHandler(device_type=devicetype, ip=deviceip, username=sshusername, password=sshpassword, secret=enablesecret)
		devicehostname = sshnet_connect.find_prompt()
		devicehostname = devicehostname.strip('#')
		if '>' in devicehostname:
			sshnet_connect.enable()
			devicehostname = devicehostname.strip('>')
			devicehostname = sshnet_connect.find_prompt()
			devicehostname = devicehostname.strip('#')
		FullOutput = sshnet_connect.send_config_set(cmdlist)
		OutputLogp = vlanfolder + '\\' + devicehost + '_log.txt'
		if os.path.exists(OutputLogp):
			OutputLog = open(OutputLogp,'a+')
		else:
			OutputLog = open(OutputLogp,'w+')
		OutputLog.write('#################################################################\n')
		OutputLog.write('Start of Configuration\n')
		OutputLog.write('Current Start Time: ' + str(datetime.now()) + '\n')
		OutputLog.write(FullOutput)
		OutputLog.write('\n')
		OutputLog.close()
		sshnet_connect.disconnect()
	except Exception as e:
		print 'Error with sending commands to ' + deviceip + '. Error is ' + str(e)
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
	except KeyboardInterrupt:
		print 'CTRL-C pressed, exiting update of switches'
		try:
			sshnet_connect.disconnect()
		except:
			'''Nothing'''
#########################################
print ''
print 'VLAN Manager'
print '############################################################'
print 'The purpose of this tool is to use a XLSX import to control'
print 'and set VLANs on various interfaces that are assigned.'
print 'Please fill in the config tab on the templated XLSX'
print 'sheet, along with all the data that you want to aaply.'
print '############################################################'
print ''
print '----Questions that need answering----'
excelfilelocation = raw_input('File to load the excel data from (e.g. C:\\Python27\\vlan-datatemplate.xlsx):')
if excelfilelocation == '':
	excelfilelocation = 'C:\\Python27\\vlan-datatemplate.xlsx'
excelfilelocation = excelfilelocation.replace('"', '')
# Load Configuration Variables
configdict = {}
for configvariables in xlhelper.sheet_to_dict(excelfilelocation,'Config'):
	try:
		configvar = configvariables.get('Variable').encode('utf-8')
		configval = configvariables.get('Value').encode('utf-8')
	except:
		configvar = configvariables.get('Variable')
		configval = configvariables.get('Value')
	configdict[configvar] = configval
# Username Variables/Questions
sshusername = configdict.get('Username')
if 'NA' == sshusername:
	sshusername = raw_input('What is the username you will use to login to the devices?:')
sshpassword = configdict.get('Password')
if 'NA' == sshpassword:
	sshpassword = getpass.getpass('What is the password you will use to login to the devices?:')
enablesecret = configdict.get('EnableSecret')
if 'NA' == enablesecret:
	enablesecret = getpass.getpass('What is the enable password you will use to access the devices?:')
# Rest of the Config Variables
databaselocation = configdict.get('DatabaseFolder')
if databaselocation == None:
	databaselocation = r'C:\Scripts\VLANManager\DB'
devicelist = []
vlanlist = []
# Create Database folder if its missing
newinstall = 0
if not os.path.exists(databaselocation):
	os.makedirs(databaselocation)
	newinstall = 1
#### VLAN DB Check (Up to date)
print 'Starting Database Check'
GetVLANs(excelfilelocation,databaselocation)
print 'Completed update of local DB'
#### Update devices with new VLAN information
continueq = raw_input('Do you want to want to update (a)ll switches or just a single (s)witch? (a/s)?:')
if 's' in continueq.lower():
	devicenameq = raw_input('What is the Hostname (in the XLSX file) of the device you want to update?:')
	for device in devicelist:
		devicehostname = device.get('Hostname').encode('utf-8')
		if devicenameq == devicehostname:
			singledevice = {}
			singledevice['Hostname'] = devicehostname
			singledevice['IP'] = device.get('IP').encode('utf-8')
			singledevice['Vendor'] = device.get('Vendor').encode('utf-8')
			singledevice['Type'] = device.get('Type').encode('utf-8')
			UpdateVLANs(singledevice,databaselocation,vlanlist)
if __name__ == "__main__":
	if 'a' in continueq.lower():
		# Start Threads
		print 'Starting update on all switches'
		for device in devicelist:	
			devicehostname = device.get('Hostname').encode('utf-8')
			deviceip = device.get('IP').encode('utf-8')
			print "Spawning Thread for " + devicehostname
			t = threading.Thread(target=UpdateVLANs, args=(device,databaselocation,vlanlist))
			t.start()
		main_thread = threading.currentThread()
		# Join All Threads
		for it_thread in threading.enumerate():
			if it_thread != main_thread:
				it_thread.join()
print 'VLAN Manager has completed updating the switches. Exiting..'
